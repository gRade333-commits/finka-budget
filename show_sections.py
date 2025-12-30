"""
Детальный анализ ВСЕХ строк листа "Сводная общая 2026г."
"""

import openpyxl
from openpyxl import load_workbook
import json

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
sheet_name = "Сводная общая 2026г."

wb = load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print(f"Всего строк: {ws.max_row}\n")

sections = {}
current_section = "header"
section_count = 0

for row_idx in range(1, min(150, ws.max_row + 1)):  # Первые 150 строк
    cell_a = ws.cell(row=row_idx, column=1).value
    cell_b = ws.cell(row=row_idx, column=2).value
    cell_c = ws.cell(row=row_idx, column=3).value
    cell_h = ws.cell(row=row_idx, column=8).value
    
    # Ищем секции (строки начинающиеся с цифры и точки)
    b_str = str(cell_b) if cell_b else ""
    
    if b_str.startswith("1.") or b_str.startswith("2.") or b_str.startswith("3.") or \
       b_str.startswith("4.") or b_str.startswith("5.") or b_str.startswith("6."):
        section_count += 1
        current_section = b_str
        print(f"\n{'='*80}")
        print(f"СЕКЦИЯ {section_count}: {b_str}")
        print(f"Строка {row_idx}")
        print(f"{'='*80}")
    
    # Показываем содержимое строк
    if any([cell_a, cell_b, cell_c, cell_h]):
        a_val = str(cell_a)[:30] if cell_a else ""
        b_val = str(cell_b)[:50] if cell_b else ""
        c_val = str(cell_c)[:20] if cell_c else ""
        h_val = str(cell_h)[:30] if cell_h else ""
        
        print(f"{row_idx:3d} | A:{a_val:30s} | B:{b_val:50s} | C:{c_val:20s} | H:{h_val}")

print(f"\n\nНайдено основных секций: {section_count}")
