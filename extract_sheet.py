"""
Извлечение конкретного листа из Excel для создания HTML-версии
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

def extract_sheet_data(file_path, sheet_name):
    """Извлекает все данные из указанного листа"""
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"Лист: {sheet_name}")
    print(f"Размер: {max_row} строк × {max_col} столбцов\n")
    
    # Собираем данные всех ячеек
    sheet_data = {
        'name': sheet_name,
        'max_row': max_row,
        'max_col': max_col,
        'cells': {},
        'merged_cells': [],
        'formulas': {}
    }
    
    # Читаем все ячейки
    for row_idx in range(1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
            
            if cell.value is not None:
                cell_info = {
                    'value': str(cell.value) if cell.value is not None else '',
                    'row': row_idx,
                    'col': col_idx,
                    'data_type': cell.data_type,
                }
                
                # Добавляем формулу если есть
                if cell.data_type == 'f':
                    cell_info['formula'] = cell.value
                    sheet_data['formulas'][cell_ref] = cell.value
                
                sheet_data['cells'][cell_ref] = cell_info
    
    # Собираем объединенные ячейки
    for merged_range in ws.merged_cells.ranges:
        sheet_data['merged_cells'].append(str(merged_range))
    
    return sheet_data

def print_sheet_structure(sheet_data):
    """Выводит структуру листа"""
    print("="*80)
    print(f"СТРУКТУРА ЛИСТА: {sheet_data['name']}")
    print("="*80)
    
    # Выводим по строкам
    for row_idx in range(1, min(sheet_data['max_row'] + 1, 25)):  # Первые 25 строк
        row_cells = []
        for col_idx in range(1, sheet_data['max_col'] + 1):
            cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
            if cell_ref in sheet_data['cells']:
                value = sheet_data['cells'][cell_ref]['value']
                if len(value) > 25:
                    value = value[:22] + "..."
                row_cells.append(value)
            else:
                row_cells.append("")
        
        # Выводим только строки с данными
        if any(row_cells):
            print(f"Строка {row_idx:2d}: {' | '.join(row_cells)}")
    
    print(f"\nОбъединенные ячейки: {len(sheet_data['merged_cells'])}")
    for merged in sheet_data['merged_cells']:
        print(f"  - {merged}")
    
    print(f"\nФормулы: {len(sheet_data['formulas'])}")
    for cell_ref, formula in list(sheet_data['formulas'].items())[:10]:
        print(f"  - {cell_ref}: {formula}")

if __name__ == "__main__":
    file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет ПУ 2026 (1).xlsx"
    sheet_name = "Свод ФОТ Алматы"
    
    data = extract_sheet_data(file_path, sheet_name)
    print_sheet_structure(data)
    
    # Сохраняем в JSON
    output_file = r"c:\Users\kiral\Desktop\finka\Новая папка\sheet_fot_almaty.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Данные сохранены в: {output_file}")
