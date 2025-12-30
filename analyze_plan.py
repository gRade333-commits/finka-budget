"""
Анализ листа План командир
"""

import openpyxl
from openpyxl import load_workbook
import json

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
sheet_name = "План командир 2026"

wb = load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print(f"Лист: {sheet_name}")
print(f"Размер: {ws.max_row} строк x {ws.max_column} столбцов\n")

# Собираем данные
data = []
for row_idx in range(1, min(50, ws.max_row + 1)):
    row_data = {"row": row_idx, "cells": {}}
    
    for col_idx in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            col_letter = chr(64 + col_idx)
            value = str(cell.value)
            if cell.data_type == 'f':
                value = f"={value}"
            row_data["cells"][col_letter] = value
    
    if row_data["cells"]:
        data.append(row_data)

# Сохраняем
with open("plan_komandir.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print("OK - сохранено в plan_komandir.json")
