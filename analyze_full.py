"""
Полный анализ всех 641 строк
"""

import openpyxl
from openpyxl import load_workbook
import json

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
sheet_name = "Сводная общая 2026г."

wb = load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print(f"Анализ листа: {sheet_name}")
print(f"Всего строк: {ws.max_row}\n")

# Собираем все данные
all_data = []
for row_idx in range(1, ws.max_row + 1):
    cell_b = ws.cell(row=row_idx, column=2)
    cell_h = ws.cell(row=row_idx, column=8)
    
    b_val = str(cell_b.value) if cell_b.value else ""
    h_val = str(cell_h.value) if cell_h.value else ""
    
    all_data.append({
        "row": row_idx,
        "name": b_val,
        "formula": h_val if cell_h.data_type == 'f' else "",
        "value": h_val if cell_h.data_type != 'f' else ""
    })

# Сохраняем в JSON
with open("svodnaya_full_data.json", "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"Данные сохранены в svodnaya_full_data.json")

# Анализ секций
sections = []
current_section = None

for item in all_data:
    b = item["name"]
    
    # Ищем основные секции
    if any([
        b.startswith("1. Фонд"),
        b.startswith("2. Приобретение питьевой"),
        b.startswith("3. Приобретение медикаментов"),
        b.startswith("4. Приобретение горюче"),
        b.startswith("5. Приобретение прочих"),
        b.startswith("6. Коммунальные"),
        "ИТОГО" in b and item["row"] > 550
    ]):
        if current_section:
            sections.append(current_section)
        current_section = {
            "name": b,
            "start": item["row"],
            "end": item["row"],
            "count": 0
        }
    
    if current_section and item["row"] > current_section["start"]:
        current_section["end"] = item["row"]
        current_section["count"] += 1

if current_section:
    sections.append(current_section)

print(f"\nНайдено основных секций: {len(sections)}\n")
for sec in sections:
    print(f"{sec['name'][:70]}")
    print(f"  Строки: {sec['start']}-{sec['end']} ({sec['count']} строк)\n")
