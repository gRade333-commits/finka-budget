"""
Быстрый анализ всех листов ПУ
"""

from openpyxl import load_workbook
import json

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет ПУ 2026 (1).xlsx"
wb = load_workbook(file_path, data_only=False)

results = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Собираем первые 20 строк
    rows_data = []
    for row_idx in range(1, min(21, ws.max_row + 1)):
        cells = {}
        for col_idx in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                col_letter = chr(64 + col_idx)
                cells[col_letter] = str(cell.value)[:50]
        if cells:
            rows_data.append({"row": row_idx, "cells": cells})
    
    results[sheet_name] = {
        "rows": ws.max_row,
        "cols": ws.max_column,
        "sample": rows_data[:10]
    }

# Сохраняем
with open("pu_all_sheets.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print("Анализ завершен!")
for name, data in results.items():
    print(f"{name}: {data['rows']} строк x {data['cols']} столбцов")
