"""
Анализ листа СС Алматы (собственные средства)
"""

import openpyxl
from openpyxl import load_workbook
import json

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет ПУ 2026 (1).xlsx"
sheet_name = "СС Алматы"

wb = load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print(f"Лист: {sheet_name}")
print(f"Размер: {ws.max_row} строк x {ws.max_column} столбцов\n")

# Собираем первые 30 строк
data = []
for row_idx in range(1, min(31, ws.max_row + 1)):
    row_data = {"row": row_idx, "cells": {}}
    
    for col_idx in range(1, min(12, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            col_letter = chr(64 + col_idx)
            value = str(cell.value)
            if cell.data_type == 'f':
                value = f"={value}"
            row_data["cells"][col_letter] = value
    
    if row_data["cells"]:
        data.append(row_data)

# Сохраняем
with open("ss_almaty.json", "w", encoding="utf-8") as f:
    json.dump({"sheet": sheet_name, "rows": ws.max_row, "cols": ws.max_column, "data": data}, f, ensure_ascii=False, indent=2)

print("OK - сохранено в ss_almaty.json")
print(f"\nПервые строки:")
for item in data[:15]:
    print(f"Строка {item['row']}: {item['cells'].get('B', '')[:60]}")
