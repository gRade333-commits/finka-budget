"""
Анализ листа СВОД ФЗП
"""

import openpyxl
from openpyxl import load_workbook

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
sheet_name = "СВОД ФЗП"

wb = load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print(f"Лист: {sheet_name}")
print(f"Размер: {ws.max_row} строк x {ws.max_column} столбцов\n")

# Анализируем первые 30 строк
for row_idx in range(1, min(31, ws.max_row + 1)):
    cells_data = []
    has_data = False
    
    for col_idx in range(1, min(15, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            has_data = True
            val = str(cell.value)
            if cell.data_type == 'f':
                val = f"={val}"
            if len(val) > 40:
                val = val[:37] + "..."
            cells_data.append(f"{chr(64+col_idx)}:{val}")
    
    if has_data:
        print(f"Строка {row_idx:2d}: {' | '.join(cells_data)}")
