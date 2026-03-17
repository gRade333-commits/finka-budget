"""
Анализ листа Калькуляция 2026
"""

import openpyxl
from openpyxl import load_workbook

file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
sheet_name = "Калькуляция 2026"

wb = load_workbook(file_path, data_only=False)
ws = wb[sheet_name]

print(f"Лист: {sheet_name}")
print(f"Размер: {ws.max_row} строк x {ws.max_column} столбцов\n")

# Анализируем первые 50 строк для понимания структуры
section_count = 0
for row_idx in range(1, min(51, ws.max_row + 1)):
    cells_data = []
    has_data = False
    
    cell_b = ws.cell(row=row_idx, column=2)
    cell_h = ws.cell(row=row_idx, column=8)
    
    b_val = str(cell_b.value) if cell_b.value else ""
    
    # Ищем заголовки разделов
    if b_val and any(x in b_val for x in ["ЗАВТРАК", "ОБЕД", "ПОЛДНИК", "УЖИН", "ВТОРОЙ УЖИН"]):
        section_count += 1
        print(f"\n{'='*70}")
        print(f"РАЗДЕЛ {section_count}: {b_val}")
        print(f"{'='*70}")
    
    # Показываем данные
    for col_idx in range(1, min(10, ws.max_column + 1)):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            has_data = True
            val = str(cell.value)
            if cell.data_type == 'f':
                val = f"={val}"
            if len(val) > 30:
                val = val[:27] + "..."
            cells_data.append(f"{chr(64+col_idx)}:{val}")
    
    if has_data:
        print(f"Строка {row_idx:3d}: {' | '.join(cells_data)}")

print(f"\n\nВсего найдено разделов питания: {section_count}")
