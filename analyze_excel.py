"""
Скрипт для анализа структуры Excel-файлов бюджета
Анализирует листы, таблицы, заголовки и структуру данных
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json
from pathlib import Path

def analyze_excel_file(file_path):
    """Анализирует структуру Excel-файла"""
    print(f"\n{'='*80}")
    print(f"АНАЛИЗ ФАЙЛА: {Path(file_path).name}")
    print(f"{'='*80}\n")
    
    # Открываем файл
    wb = load_workbook(file_path, data_only=False)
    
    file_structure = {
        'filename': Path(file_path).name,
        'total_sheets': len(wb.sheetnames),
        'sheets': []
    }
    
    # Анализируем каждый лист
    for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"\n📄 ЛИСТ {sheet_idx}: {sheet_name}")
        print("-" * 80)
        
        ws = wb[sheet_name]
        
        # Получаем размеры листа
        max_row = ws.max_row
        max_col = ws.max_column
        
        print(f"   Размер: {max_row} строк × {max_col} столбцов")
        
        # Анализируем первые строки (обычно там заголовки)
        print(f"\n   Первые 10 строк:")
        
        sheet_data = {
            'name': sheet_name,
            'index': sheet_idx,
            'max_row': max_row,
            'max_column': max_col,
            'headers': [],
            'sample_data': []
        }
        
        # Читаем первые 10 строк для анализа структуры
        for row_idx in range(1, min(11, max_row + 1)):
            row_data = []
            row_text = f"   Строка {row_idx:2d}: "
            
            for col_idx in range(1, min(max_col + 1, 20)):  # Ограничиваем 20 колонками для вывода
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # Сокращаем длинные значения
                if value is not None:
                    str_value = str(value)
                    if len(str_value) > 30:
                        str_value = str_value[:27] + "..."
                    row_data.append(str_value)
                else:
                    row_data.append("")
            
            sheet_data['sample_data'].append(row_data)
            
            # Выводим первые 5 значимых ячеек
            non_empty = [v for v in row_data if v]
            if non_empty:
                print(row_text + " | ".join(non_empty[:5]))
        
        # Проверяем наличие объединенных ячеек
        if ws.merged_cells:
            print(f"\n   ⚠️ Объединенные ячейки: {len(ws.merged_cells.ranges)} групп")
            for merged_range in list(ws.merged_cells.ranges)[:5]:  # Показываем первые 5
                print(f"      • {merged_range}")
        
        # Проверяем формулы
        formulas = []
        for row in ws.iter_rows(min_row=1, max_row=min(20, max_row)):
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formulas.append(f"{cell.coordinate}: {cell.value}")
        
        if formulas:
            print(f"\n   📐 Найдено формул: {len(formulas)}")
            for formula in formulas[:3]:  # Показываем первые 3
                print(f"      • {formula}")
        
        file_structure['sheets'].append(sheet_data)
        print()
    
    return file_structure

def save_structure_to_json(structure, output_file):
    """Сохраняет структуру в JSON файл"""
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(structure, f, ensure_ascii=False, indent=2)
    print(f"\n✅ Структура сохранена в: {output_file}")

def main():
    # Пути к файлам
    budget_rb = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
    budget_pu = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет ПУ 2026 (1).xlsx"
    
    structures = {}
    
    # Анализируем Республиканский бюджет
    try:
        structures['budget_rb'] = analyze_excel_file(budget_rb)
    except Exception as e:
        print(f"\n❌ Ошибка при анализе РБ: {e}")
    
    # Анализируем Собственные средства
    try:
        structures['budget_pu'] = analyze_excel_file(budget_pu)
    except Exception as e:
        print(f"\n❌ Ошибка при анализе ПУ: {e}")
    
    # Сохраняем общую структуру
    output_json = r"c:\Users\kiral\Desktop\finka\Новая папка\excel_structure.json"
    save_structure_to_json(structures, output_json)
    
    # Выводим итоговую сводку
    print(f"\n{'='*80}")
    print("ИТОГОВАЯ СВОДКА")
    print(f"{'='*80}\n")
    
    for budget_type, structure in structures.items():
        if structure:
            print(f"📊 {structure['filename']}")
            print(f"   Всего листов: {structure['total_sheets']}")
            print(f"   Листы: {', '.join([s['name'] for s in structure['sheets']])}")
            print()
    
    print("\n✨ Анализ завершён!")
    print(f"📁 Результаты сохранены в: excel_structure.json")
    print("\nСледующий шаг: изучите структуру и выберите листы для реализации в первую очередь")

if __name__ == "__main__":
    main()
