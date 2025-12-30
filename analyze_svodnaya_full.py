"""
Полный анализ листа "Сводная общая 2026г." с сохранением в файл
"""

import openpyxl
from openpyxl import load_workbook
import json

def analyze_full_structure(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    result = {
        "sheet": sheet_name,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "sections": []
    }
    
    current_section = None
    section_start = 0
    
    for row_idx in range(1, ws.max_row + 1):
        cell_b = ws.cell(row=row_idx, column=2)  # Столбец B - наименование
        cell_h = ws.cell(row=row_idx, column=8)  # Столбец H - сумма
        
        value_b = str(cell_b.value) if cell_b.value else ""
        
        # Определяем начало новой секции (жирный текст, номер с точкой)
        if value_b and ("1. Фонд" in value_b or 
                       "2. Приобретение питьевой" in value_b or
                       "3. Приобретение медикаментов" in value_b or
                       "4. Приобретение горюче" in value_b or
                       "5. Приобретение прочих" in value_b or
                       "6. Коммунальные" in value_b or
                       "ИТОГО" in value_b):
            
            if current_section:
                current_section["end_row"] = row_idx - 1
                current_section["row_count"] = current_section["end_row"] - current_section["start_row"] + 1
                result["sections"].append(current_section)
            
            current_section = {
                "name": value_b,
                "start_row": row_idx,
                "end_row": 0,
                "row_count": 0,
                "items": []
            }
        
        # Собираем элементы секции
        if current_section and value_b and row_idx > current_section["start_row"]:
            formula = ""
            if cell_h.data_type == 'f':
                formula = str(cell_h.value)
            
            current_section["items"].append({
                "row": row_idx,
                "name": value_b,
                "formula": formula
            })
    
    # Закрываем последнюю секцию
    if current_section:
        current_section["end_row"] = ws.max_row
        current_section["row_count"] = current_section["end_row"] - current_section["start_row"] + 1
        result["sections"].append(current_section)
    
    # Сохраняем в файл
    with open("svodnaya_structure.json", "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)
    
    print(f"Лист: {sheet_name}")
    print(f"Всего строк: {ws.max_row}, столбцов: {ws.max_column}")
    print(f"\nНайдено секций: {len(result['sections'])}")
    print("\nСтруктура секций:")
    for section in result["sections"]:
        print(f"\n{section['name']}")
        print(f"  Строки: {section['start_row']}-{section['end_row']} ({section['row_count']} строк)")
        print(f"  Элементов: {len(section['items'])}")
        if len(section['items']) > 0:
            print(f"  Первые 3 элемента:")
            for item in section['items'][:3]:
                print(f"    - Строка {item['row']}: {item['name'][:60]}")
    
    print(f"\nДетали сохранены в: svodnaya_structure.json")
    return result

if __name__ == "__main__":
    file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
    sheet_name = "Сводная общая 2026г."
    
    analyze_full_structure(file_path, sheet_name)
