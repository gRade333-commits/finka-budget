"""
Извлечение листа "ДОХОДНАЯ ЧАСТЬ РБ"
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def extract_sheet_detailed(file_path, sheet_name):
    wb = load_workbook(file_path, data_only=False)
    ws = wb[sheet_name]
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"\n{'='*100}")
    print(f"ЛИСТ: {sheet_name}")
    print(f"Размер: {max_row} строк × {max_col} столбцов")
    print(f"{'='*100}\n")
    
    for row_idx in range(1, max_row + 1):
        row_cells = []
        has_data = False
        
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            if value is not None:
                has_data = True
                str_value = str(value)
                if cell.data_type == 'f':
                    str_value = f"={value}"
                if len(str_value) > 50:
                    str_value = str_value[:47] + "..."
                row_cells.append((get_column_letter(col_idx), str_value))
        
        if has_data:
            print(f"Строка {row_idx:2d}:")
            for col, val in row_cells:
                print(f"  {col}: {val}")
            print()
    
    print(f"\nОбъединенные ячейки ({len(ws.merged_cells.ranges)}):")
    for merged in ws.merged_cells.ranges:
        print(f"  {merged}")

if __name__ == "__main__":
    file_path = r"c:\Users\kiral\Desktop\finka\Новая папка\Бюджет РБ 2026 (1).xlsx"
    sheet_name = "ДОХОДНАЯ ЧАСТЬ РБ"
    
    extract_sheet_detailed(file_path, sheet_name)
